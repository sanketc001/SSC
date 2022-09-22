"""
Superclass - adding data to Grade Sheet method
"""
import pandas
import openpyxl
import openpyxl.utils.dataframe
import datetime
import os
import sys

class GradeSheetPrintSSC():
    def __init__(self):
        self.gradesheetprinterprimer = {}
        self.gradesheetprinter = {}
        self.permpathtoexcelssc = r'C:\SSC\SimpleStockChecker_REV1\sscpackage\storage\excelstorage\\'
        self.instancepath = ""
        self.logfilename = ""

    def setsheetnamesscgr(self, ticker):
        curdatessc = datetime.datetime.today().strftime("%d_%m_%y")
        self.sheetnamesscgr = ticker + "__" + curdatessc

    def setinstancepath(self, ticker, uniquerunid):
        self.instancepath = self.permpathtoexcelssc + str(ticker) + "__" + str(uniquerunid) + '.xlsx'

    def create_pathinit(self):
        try:
            if os.path.exists(self.permpathtoexcelssc):
                return 1
            else:
                os.makedirs(self.permpathtoexcelssc)
        except Exception as er:
            print(f'{self.__name__} - Error at create_pathinit')
            print(f'This: {er.__class__} occurred')


    def printprimer(self, sectionname, ticker, uniquerunid, sectorssc, industryssc):
        self.create_pathinit()
        curdatessc = datetime.datetime.today().strftime("%d_%m_%y")
        self.gradesheetprinterprimer.update({"Ticker": str(ticker), "Unique Run ID": str(uniquerunid),
                                             "Sector": str(sectorssc), "Industry": str(industryssc),
                                             "Date / Time": datetime.datetime.today().strftime("%d-%m-%y  %H:%M:%S"),
                                             "Section": str(sectionname)})
        localdfssc = pandas.DataFrame.from_dict(self.gradesheetprinterprimer, orient='index')
        self.logfilename = str(ticker) + '__' + str(uniquerunid) + ".xlsx"
        self.instancepath = str(self.permpathtoexcelssc) + str(self.logfilename)
        self.sheetnamesscgr = str(ticker) + '__' + str(curdatessc)
        with pandas.ExcelWriter(path=self.instancepath, engine='xlsxwriter', mode='w') as exwssc:
            localdfssc.to_excel(exwssc, sheet_name=(str(ticker) + '__' + str(curdatessc)))

    def gradeprinterssc(self, **data):
        self.gradesheetprinter.update(**data)
        # TODO: flesh out printing data to excel and learn more about numpy/pandas its NUMPIE not frunumpy

    def sectionprinttoexcel(self):
        sectiondf = pandas.DataFrame.from_dict(self.gradesheetprinter, orient="index")
        wbsscgr = openpyxl.load_workbook(self.instancepath)
        wssscgr = wbsscgr[self.sheetnamesscgr]
        for rowitem in openpyxl.utils.dataframe.dataframe_to_rows(sectiondf, index=True, header=False):
            wssscgr.append(rowitem)

        wbsscgr.save(self.instancepath)

    def sectionendprinttoexcel(self, **kwargs):
        localpoint = 0
        localtotal = 0
        localdict = kwargs
        for key in localdict.keys():
            localpoint += localdict[key]["Current Points"]
            localtotal += localdict[key]["Base Points"]

        statementstring = {"TOTAL CURRENT POINTS": localpoint, "TOTAL POSSIBLE POINTS": localtotal}

        sectionend = pandas.DataFrame.from_dict(statementstring, orient="index")
        sectionend.convert_dtypes()
        wb = openpyxl.load_workbook(self.instancepath)
        wssscgr = wb[self.sheetnamesscgr]
        for rowitem in openpyxl.utils.dataframe.dataframe_to_rows(sectionend, index=True, header=False):
            wssscgr.append(rowitem)

        wb.save(self.instancepath)

        return localpoint, localtotal


